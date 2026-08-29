# /// script
# dependencies = [
#     "streamlit>=1.30.0",
#     "playwright>=1.40.0",
#     "pandas>=2.0.0",
#     "openpyxl>=3.1.0",
#     "python-dateutil>=2.8.2",
# ]
# ///

import os
import re
import sys
import time
from datetime import datetime, timedelta
import pandas as pd
import streamlit as st

# Set Streamlit Page Configuration for Mobile & Desktop
st.set_page_config(
    page_title="CGNET Employee Audit Portal",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Custom CSS for Modern, WOW Aesthetics & Mobile Responsiveness
st.markdown("""
<style>
    @import url('https://fonts.googleapis.com/css2?family=DM+Sans:wght@400;500;700&display=swap');
    
    html, body, [class*="css"] {
        font-family: 'DM Sans', sans-serif;
    }
    
    .stApp {
        background: linear-gradient(135deg, #0f172a 0%, #1e293b 100%);
        color: #f8fafc;
    }
    
    /* Header Container */
    .main-header {
        background: linear-gradient(90deg, #1e3a8a 0%, #3b82f6 100%);
        padding: 1.5rem 2rem;
        border-radius: 12px;
        box-shadow: 0 10px 25px -5px rgba(59, 130, 246, 0.3);
        margin-bottom: 2rem;
        text-align: center;
    }
    .main-header h1 {
        color: #ffffff !important;
        font-weight: 700;
        margin: 0;
        font-size: 2.2rem;
    }
    .main-header p {
        color: #e0f2fe !important;
        margin-top: 0.5rem;
        font-size: 1rem;
    }
    
    /* KPI Card Container */
    .kpi-card {
        background: rgba(30, 41, 59, 0.7);
        border: 1px solid rgba(255, 255, 255, 0.1);
        backdrop-filter: blur(10px);
        padding: 1.2rem;
        border-radius: 12px;
        text-align: center;
        transition: transform 0.2s ease, box-shadow 0.2s ease;
    }
    .kpi-card:hover {
        transform: translateY(-3px);
        box-shadow: 0 8px 20px rgba(0, 0, 0, 0.3);
    }
    .kpi-title {
        color: #94a3b8;
        font-size: 0.9rem;
        text-transform: uppercase;
        letter-spacing: 1px;
        margin-bottom: 0.5rem;
    }
    .kpi-value {
        color: #38bdf8;
        font-size: 1.8rem;
        font-weight: 700;
    }
    .kpi-subtext {
        color: #64748b;
        font-size: 0.8rem;
        margin-top: 0.3rem;
    }
    
    /* Download Button */
    .stDownloadButton > button {
        background: linear-gradient(90deg, #10b981 0%, #059669 100%) !important;
        color: white !important;
        font-weight: 700 !important;
        border: none !important;
        padding: 0.75rem 2rem !important;
        border-radius: 8px !important;
        box-shadow: 0 4px 14px rgba(16, 185, 129, 0.4) !important;
        width: 100%;
    }
    
    /* Status Box */
    .status-box {
        background: #1e293b;
        border-left: 4px solid #3b82f6;
        padding: 1rem;
        border-radius: 4px;
        font-family: monospace;
        max-height: 200px;
        overflow-y: auto;
        font-size: 0.85rem;
    }
</style>
""", unsafe_allow_html=True)

# Auto-ensure Playwright chromium binary is installed in cloud environment
import subprocess
@st.cache_resource
def setup_playwright():
    try:
        subprocess.run([sys.executable, "-m", "playwright", "install", "chromium"], check=False)
    except Exception:
        pass
    return True

setup_playwright()

# Import scraper from audit_automation module
import audit_automation

# Title Banner
st.markdown("""
<div class="main-header">
    <h1>⚡ CGNET Automated Employee Audit Portal</h1>
    <p>Run real-time performance audit scraping, analyze ticket resolutions & download executive reports from any device.</p>
</div>
""", unsafe_allow_html=True)

# Sidebar Inputs
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Sidebar Inputs
with st.sidebar:
    st.header("📋 Audit Configuration")
    
    # Employee Selection
    emp_options = [
        "👥 ALL TEAM (Ajit, Shashikant, Om, Sabin, Sunil, Sanjeev, Chandramani)",
        "Ajit Shrestha",
        "Shashikant Chaudhary",
        "Om Neupane",
        "Sabin Giri",
        "Sunil Chaudhary",
        "Sanjeev Giri",
        "Chandramani Tharu",
        "Custom Employee Name / List..."
    ]
    selected_emp_type = st.selectbox("Employee / Team Selection", emp_options, index=0)
    
    if selected_emp_type == "Custom Employee Name / List...":
        employee_name = st.text_input("Enter Employee Name(s) (comma separated for multiple)", value="Om Neupane")
    elif selected_emp_type == "👥 ALL TEAM (Ajit, Shashikant, Om, Sabin, Sunil, Sanjeev, Chandramani)":
        employee_name = "Ajit Shrestha, Shashikant Chaudhary, Om Neupane, Sabin Giri, Sunil Chaudhary, Sanjeev Giri, Chandramani Tharu"
    else:
        employee_name = selected_emp_type
        
    st.markdown("---")
    st.subheader("📅 Date Range Selector")
    
    today = datetime.now().date()
    yesterday = today - timedelta(days=1)
    
    date_preset = st.radio("Quick Range", ["Today", "Yesterday & Today (2 Days)", "Last 7 Days", "Custom Range"], index=1)
    
    if date_preset == "Today":
        from_date_obj = today
        to_date_obj = today
    elif date_preset == "Yesterday & Today (2 Days)":
        from_date_obj = yesterday
        to_date_obj = today
    elif date_preset == "Last 7 Days":
        from_date_obj = today - timedelta(days=6)
        to_date_obj = today
    else:
        col_d1, col_d2 = st.columns(2)
        with col_d1:
            from_date_obj = st.date_input("From Date", value=yesterday)
        with col_d2:
            to_date_obj = st.date_input("To Date", value=today)

    from_date_str = from_date_obj.strftime("%d %b %Y")  # e.g., "04 Aug 2026"
    to_date_str = to_date_obj.strftime("%d %b %Y")      # e.g., "05 Aug 2026"

    st.markdown(f"**Target Period:** `{from_date_str}` to `{to_date_str}`")
    st.markdown("---")
    
    run_btn = st.button("🚀 Run Audit Scraper", type="primary", use_container_width=True)

# Helper function to generate Master Executive Excel Report from multiple DataFrames
def build_executive_team_excel(master_df, team_summary_df, cat_summary_df):
    output_buf = io.BytesIO()
    with pd.ExcelWriter(output_buf, engine='openpyxl') as writer:
        team_summary_df.to_excel(writer, sheet_name="Team Executive Summary", index=False)
        cat_summary_df.to_excel(writer, sheet_name="Category Summary", index=False)
        master_df.to_excel(writer, sheet_name="Master Audit Details", index=False)
        
    output_buf.seek(0)
    wb = openpyxl.load_workbook(output_buf)
    
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    cell_font = Font(name="Segoe UI", size=10)
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'), bottom=Side(style='thin', color='D9D9D9')
    )
    
    for sheetname in wb.sheetnames:
        ws = wb[sheetname]
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
        for row in range(2, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                cell.font = cell_font
                cell.border = thin_border
                if sheetname == "Team Executive Summary" and row == ws.max_row:
                    cell.font = Font(name="Segoe UI", size=11, bold=True, color="1F4E78")
                    cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
                    
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 45)
            
    final_buf = io.BytesIO()
    wb.save(final_buf)
    final_buf.seek(0)
    return final_buf.getvalue()

def classify_work_type(row):
    remark = str(row.get('Grid Remark', '')).strip()
    solution = str(row.get('Solution Given', '')).strip()
    employee_note = str(row.get('Employee Remark / Solution Note', '')).strip()
    title = str(row.get('Title', '')).strip()
    category = str(row.get('Category', '')).strip()
    
    full_text = f"{title} {remark} {solution} {employee_note}".upper()
    
    # Check for 'ALCLFE' (WiFi 6 Serial Number prefix) or Old/New SN router upgrade remarks
    is_wifi6_upgrade = (
        'ALCLFE' in full_text or
        ('WIFI 6' in full_text and ('OLD' in full_text or 'NEW' in full_text or 'UPGRADE' in full_text or 'UPGARDE' in full_text or 'SN' in full_text)) or
        ('WIFI' in full_text and 'ALCL' in full_text)
    )
    
    if is_wifi6_upgrade:
        return 'WiFi 6 Upgrade'
    elif 'IPTV' in full_text or 'IPTV' in category.upper():
        return 'IPTV Issue'
    else:
        return 'General / Other Issues'

# Mode Selection Tabs in Main Area
main_mode_tab1, main_mode_tab2 = st.tabs(["📊 Performance Analytics & Scraper", "📁 Combine Uploaded Reports (Manager Tool)"])

with main_mode_tab1:
    if run_btn:
        st.info(f"⏳ Starting automated scraper for **{employee_name}** from `{from_date_str}` to `{to_date_str}`...")
        
        log_container = st.empty()
        logs_list = []
        
        def ui_log(msg, level="INFO"):
            timestamp = datetime.now().strftime("%H:%M:%S")
            log_line = f"[{timestamp}] [{level}] {msg}"
            logs_list.append(log_line)
            log_html = "<br>".join(logs_list[-8:])
            log_container.markdown(f'<div class="status-box">{log_html}</div>', unsafe_allow_html=True)

        progress_bar = st.progress(0.1, text="Initializing Playwright browser...")

        try:
            sys.argv = [
                "audit_automation.py",
                "--employee", employee_name,
                "--from-date", from_date_str,
                "--to-date", to_date_str,
                "--non-interactive"
            ]
            
            import subprocess
            try:
                subprocess.run([sys.executable, "-m", "playwright", "install", "chromium"], check=False)
            except Exception:
                pass

            audit_automation.main()
            
            progress_bar.progress(1.0, text="Scraping completed!")
            st.success("✅ Audit Scraper completed successfully!")
            
            if ',' in employee_name or 'ALL TEAM' in employee_name.upper():
                safe_emp = "ALL_TEAM"
            else:
                safe_emp = re.sub(r'[^a-zA-Z0-9]', '_', employee_name)
            safe_from = re.sub(r'[^a-zA-Z0-9]', '_', from_date_str)
            output_file = f"audit_report_{safe_emp}_{safe_from}.xlsx"
            
            if os.path.exists(output_file):
                st.session_state["last_output_file"] = output_file
                st.session_state["last_run_time"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            else:
                import glob
                matching_files = glob.glob(f"audit_report_*{safe_from}*.xlsx")
                if matching_files:
                    st.session_state["last_output_file"] = matching_files[0]
                    st.session_state["last_run_time"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                
        except Exception as e:
            st.error(f"❌ Scraper error encountered: {e}")
            st.exception(e)

    # Display Dashboard if output file exists
    output_file = st.session_state.get("last_output_file", None)

    if output_file and os.path.exists(output_file):
        st.markdown("### 📊 Performance Analytics Dashboard")
        st.caption(f"Last updated: {st.session_state.get('last_run_time', 'Recently')} | Report File: `{os.path.basename(output_file)}`")
        
        try:
            xls = pd.ExcelFile(output_file)
            df_details = pd.read_excel(xls, sheet_name="Audit Details")
            df_summary = pd.read_excel(xls, sheet_name="Summary Report", nrows=5)
            
            df_details['Task / Issue Type'] = df_details.apply(classify_work_type, axis=1)
            
            col1, col2, col3, col4, col5 = st.columns(5)
            
            def get_val(metric_name):
                row = df_summary[df_summary["Metric"].str.contains(metric_name, case=False, na=False)]
                if not row.empty:
                    return str(row.iloc[0]["Value"])
                return "N/A"

            with col1:
                st.markdown(f"""
                <div class="kpi-card">
                    <div class="kpi-title">Total Records</div>
                    <div class="kpi-value">{get_val("Total Audit Records Scraped")}</div>
                    <div class="kpi-subtext">Scraped logs</div>
                </div>
                """, unsafe_allow_html=True)
                
            with col2:
                st.markdown(f"""
                <div class="kpi-card">
                    <div class="kpi-title">Solved / Handled</div>
                    <div class="kpi-value">{get_val("Total Tickets Solved")}</div>
                    <div class="kpi-subtext">Completed or MS Assigned</div>
                </div>
                """, unsafe_allow_html=True)
                
            with col3:
                st.markdown(f"""
                <div class="kpi-card">
                    <div class="kpi-title">Solution Rate</div>
                    <div class="kpi-value">{get_val("Ticket Solution Rate")}</div>
                    <div class="kpi-subtext">Completion %</div>
                </div>
                """, unsafe_allow_html=True)
                
            with col4:
                st.markdown(f"""
                <div class="kpi-card">
                    <div class="kpi-title">Avg Time (Assigned)</div>
                    <div class="kpi-value" style="font-size:1.4rem;">{get_val("Average Completion Time (From Assigned Date)")}</div>
                    <div class="kpi-subtext">From assigned timestamp</div>
                </div>
                """, unsafe_allow_html=True)
                
            with col5:
                st.markdown(f"""
                <div class="kpi-card">
                    <div class="kpi-title">Avg Time (Created)</div>
                    <div class="kpi-value" style="font-size:1.4rem;">{get_val("Average Completion Time (From Created Date)")}</div>
                    <div class="kpi-subtext">From initial creation</div>
                </div>
                """, unsafe_allow_html=True)

            st.markdown("<br>", unsafe_allow_html=True)

            st.markdown("### 📥 Download Executive Report")
            with open(output_file, "rb") as f:
                bytes_data = f.read()
                st.download_button(
                    label=f"⬇️ Download Excel Audit Report ({os.path.basename(output_file)})",
                    data=bytes_data,
                    file_name=os.path.basename(output_file),
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

            st.markdown("<br>", unsafe_allow_html=True)
            
            tab1, tab2, tab3 = st.tabs(["📝 Scraped Audit Details", "🏷️ Task & Category Breakdown", "📊 Raw Summary Sheet"])
            
            with tab1:
                st.subheader("Filterable Audit Details Table")
                search_query = st.text_input("🔍 Search records by ticket #, remark, or account name...", "")
                
                if search_query:
                    filtered_df = df_details[df_details.astype(str).apply(lambda row: row.str.contains(search_query, case=False).any(), axis=1)]
                else:
                    filtered_df = df_details
                    
                st.dataframe(filtered_df, use_container_width=True, height=400)
                
            with tab2:
                st.subheader("Specific Task / Issue Breakdown (WiFi 6, IPTV, Router, etc.)")
                work_counts = df_details['Task / Issue Type'].value_counts().reset_index()
                work_counts.columns = ['Task / Issue Type', 'Total Tickets Count']
                work_counts['% Share'] = (work_counts['Total Tickets Count'] / len(df_details) * 100).round(1).astype(str) + '%'
                
                st.dataframe(work_counts, use_container_width=True)
                st.bar_chart(work_counts.set_index('Task / Issue Type')['Total Tickets Count'])
                
                st.markdown("---")
                st.subheader("Portal Category & Sub-Category Breakdown")
                if 'Category' in df_details.columns:
                    sub_col = 'Sub Category' if 'Sub Category' in df_details.columns else ('Sub Sub Category' if 'Sub Sub Category' in df_details.columns else None)
                    group_cols = ['Category'] + ([sub_col] if sub_col else [])
                    
                    cat_counts = df_details.groupby(group_cols, dropna=False).size().reset_index(name='Total Tickets Count')
                    cat_counts = cat_counts.sort_values(by='Total Tickets Count', ascending=False)
                    
                    st.dataframe(cat_counts, use_container_width=True)
                else:
                    st.info("No Category column found in details dataset.")
                    
            with tab3:
                st.subheader("Summary Report Sheet View")
                df_full_summary = pd.read_excel(xls, sheet_name="Summary Report")
                st.dataframe(df_full_summary, use_container_width=True)

        except Exception as read_err:
            st.error(f"Could not load output preview: {read_err}")
    else:
        st.info("👈 Select an Employee or **👥 ALL TEAM**, set Date Range, then click **Run Audit Scraper** to generate your report.")

with main_mode_tab2:
    st.subheader("📁 Executive Team Report Merger (Manager Tool)")
    st.write("Upload individual employee audit Excel files (`audit_report_*.xlsx`) to merge them into a single **Executive Team Report** for management!")
    
    uploaded_files = st.file_uploader(
        "Upload Individual Employee Audit Excel Reports",
        type=["xlsx"],
        accept_multiple_files=True,
        help="Select multiple files (e.g. Om Neupane, Ajit Shrestha, Shashikant Chaudhary, Sabin Giri, Sunil Chaudhary, Sanjeev Giri, Chandramani Tharu, etc.)"
    )
    
    if uploaded_files:
        st.success(f"📥 Received {len(uploaded_files)} Excel report file(s) for team merging.")
        
        all_dfs = []
        for file in uploaded_files:
            try:
                xls = pd.ExcelFile(file)
                if "Audit Details" in xls.sheet_names:
                    df = pd.read_excel(xls, sheet_name="Audit Details")
                    
                    emp_name = None
                    if "Grid Employee Name" in df.columns and not df["Grid Employee Name"].dropna().empty:
                        emp_name = df["Grid Employee Name"].dropna().iloc[0]
                    else:
                        match = re.search(r'audit_report_(.+?)_\d', file.name)
                        if match:
                            emp_name = match.group(1).replace('_', ' ')
                        else:
                            emp_name = file.name.replace('.xlsx', '')
                            
                    df["Employee Name"] = emp_name
                    df["Task / Issue Type"] = df.apply(classify_work_type, axis=1)
                    all_dfs.append(df)
            except Exception as read_err:
                st.warning(f"Could not read {file.name}: {read_err}")
                
        if all_dfs:
            master_df = pd.concat(all_dfs, ignore_index=True)
            
            def is_solved(row):
                st_val = str(row.get("Status", "")).strip().lower()
                rm_val = str(row.get("Grid Remark", "")).strip().lower()
                if st_val in ["completed", "closed"]:
                    return True
                if "ms" in rm_val or "assign" in rm_val or "transfer" in rm_val or "forward" in rm_val:
                    return True
                return False

            master_df["Is_Solved_Val"] = master_df.apply(is_solved, axis=1)
            
            # Team Summary Calculation
            summary_rows = []
            for emp, grp in master_df.groupby("Employee Name"):
                tot = len(grp)
                solved = grp["Is_Solved_Val"].sum()
                rate = f"{(solved / tot * 100):.1f}%" if tot > 0 else "0.0%"
                summary_rows.append({
                    "Employee Name": emp,
                    "Total Scraped Tickets": tot,
                    "Solved / Handled Count": solved,
                    "Solution Rate %": rate
                })
                
            team_summary_df = pd.DataFrame(summary_rows)
            
            # Total Row
            tot_tickets_team = len(master_df)
            tot_solved_team = master_df["Is_Solved_Val"].sum()
            team_rate_val = f"{(tot_solved_team / tot_tickets_team * 100):.1f}%" if tot_tickets_team > 0 else "0.0%"
            
            total_team_row = pd.DataFrame([{
                "Employee Name": "👥 GRAND TOTAL (ALL TEAM)",
                "Total Scraped Tickets": tot_tickets_team,
                "Solved / Handled Count": tot_solved_team,
                "Solution Rate %": team_rate_val
            }])
            team_summary_df = pd.concat([team_summary_df, total_team_row], ignore_index=True)
            
            # Task / Issue Breakdown Calculation
            work_summary_df = master_df.groupby("Task / Issue Type", dropna=False).agg(
                Total_Tickets=("Ticket Number", "count"),
                Solved_Count=("Is_Solved_Val", "sum")
            ).reset_index()
            work_summary_df["Solution Rate %"] = (work_summary_df["Solved_Count"] / work_summary_df["Total_Tickets"] * 100).round(1).astype(str) + '%'
            work_summary_df["% Share of Total"] = (work_summary_df["Total_Tickets"] / len(master_df) * 100).round(1).astype(str) + '%'
            
            # Category Summary Calculation
            cat_summary_df = master_df.groupby("Category", dropna=False).agg(
                Total_Tickets=("Ticket Number", "count"),
                Solved_Count=("Is_Solved_Val", "sum")
            ).reset_index()
            cat_summary_df["Solution Rate %"] = (cat_summary_df["Solved_Count"] / cat_summary_df["Total_Tickets"] * 100).round(1).astype(str) + '%'
            
            export_master = master_df.drop(columns=["Is_Solved_Val"], errors="ignore")
            
            st.markdown("### 🏆 Executive Team Performance Summary")
            
            c1, c2, c3 = st.columns(3)
            with c1:
                st.metric("Total Team Tickets Scraped", tot_tickets_team)
            with c2:
                st.metric("Total Team Solved / Handled", tot_solved_team)
            with c3:
                st.metric("Overall Team Solution Rate", team_rate_val)
                
            st.markdown("#### Employee Workload & Performance Comparison")
            st.dataframe(team_summary_df, use_container_width=True)
            
            st.markdown("---")
            st.markdown("### 🏷️ Specific Task & Issue Breakdown (WiFi 6 Upgrades, IPTV, Hardware, etc.)")
            st.dataframe(work_summary_df, use_container_width=True)
            st.bar_chart(work_summary_df.set_index("Task / Issue Type")["Total_Tickets"])
            
            st.markdown("#### Tickets Workload per Employee")
            st.bar_chart(team_summary_df[team_summary_df["Employee Name"] != "👥 GRAND TOTAL (ALL TEAM)"].set_index("Employee Name")["Total Scraped Tickets"])
            
            st.markdown("#### Combined Filterable Master Audit Table")
            st.dataframe(export_master, use_container_width=True, height=400)
            
            excel_bytes = build_executive_team_excel(export_master, team_summary_df, work_summary_df)
            today_filename_str = datetime.now().strftime("%d_%b_%Y")
            exec_file_name = f"EXECUTIVE_TEAM_AUDIT_REPORT_{today_filename_str}.xlsx"
            
            st.download_button(
                label=f"⬇️ Download Combined Executive Team Report ({exec_file_name})",
                data=excel_bytes,
                file_name=exec_file_name,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
