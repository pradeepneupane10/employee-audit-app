# /// script
# dependencies = [
#     "playwright>=1.40.0",
#     "pandas>=2.0.0",
#     "openpyxl>=3.1.0",
#     "python-dateutil>=2.8.2",
# ]
# ///

import os
import re
import sys
import argparse
from datetime import datetime, timedelta
import pandas as pd
from dateutil import parser as date_parser
from playwright.sync_api import sync_playwright

# Credentials
LOGIN_USER = "laxman.koirala"
LOGIN_PASS = "koirala...laxman"

# Setup simple, clean logging
def log(msg, level="INFO"):
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    print(f"[{timestamp}] [{level}] {msg}")

def wait_for_postback(page, timeout_ms=8000):
    try:
        page.evaluate("""
        () => {
            return new Promise((resolve) => {
                if (typeof Sys === 'undefined' || !Sys.WebForms || !Sys.WebForms.PageRequestManager) {
                    resolve();
                    return;
                }
                const prm = Sys.WebForms.PageRequestManager.getInstance();
                if (!prm.get_isInAsyncPostBack()) {
                    resolve();
                    return;
                }
                const handler = () => {
                    if (!prm.get_isInAsyncPostBack()) {
                        prm.remove_endRequest(handler);
                        resolve();
                    }
                };
                prm.add_endRequest(handler);
                setTimeout(() => {
                    prm.remove_endRequest(handler);
                    resolve();
                }, 7500);
            });
        }
        """)
    except Exception:
        page.wait_for_timeout(500)

def safe_wait_for_networkidle(page, timeout_ms=5000):
    try:
        page.wait_for_load_state("networkidle", timeout=timeout_ms)
    except Exception:
        pass

def format_duration(td):
    if pd.isna(td) or not isinstance(td, timedelta):
        return "N/A"
    total_seconds = int(td.total_seconds())
    if total_seconds < 0:
        return "Negative Time"
    days = total_seconds // 86400
    hours = (total_seconds % 86400) // 3600
    minutes = (total_seconds % 3600) // 60
    
    parts = []
    if days > 0:
        parts.append(f"{days}d")
    if hours > 0:
        parts.append(f"{hours}h")
    if minutes > 0 or not parts:
        parts.append(f"{minutes}m")
    return " ".join(parts)

def parse_date_robustly(date_str):
    if not date_str or pd.isna(date_str):
        return None
    # Normalize whitespaces
    date_str = re.sub(r'\s+', ' ', str(date_str).strip())
    if not date_str:
        return None
    
    # Try dateutil parser first as it is very smart
    try:
        return date_parser.parse(date_str)
    except Exception:
        pass
    
    # Fallback to standard formats
    formats = [
        "%d %b %Y %I:%M:%S %p",  # 05 Aug 2026 05:04:12 PM
        "%d %b %Y %I:%M %p",     # 05 Aug 2026 05:04 PM
        "%d-%b-%Y %I:%M:%S %p",
        "%d-%b-%Y %I:%M %p",
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d %I:%M:%S %p",
    ]
    for fmt in formats:
        try:
            return datetime.strptime(date_str, fmt)
        except ValueError:
            continue
    return None

def handle_login(page):
    try:
        page.locator('input[type="password"]').first.wait_for(state="visible", timeout=6000)
    except Exception as e:
        log(f"Password field wait failed or timed out: {e}. Checking if logged in.")
        # Take a screenshot to help diagnose page state
        try:
            page.screenshot(path="login_error_debug.png")
            log("Saved debug screenshot to login_error_debug.png")
        except Exception as screenshot_err:
            log(f"Failed to capture login screenshot: {screenshot_err}")
        return
        
    log("Login page detected. Performing automatic login...")
    
    # Find username input
    username_field = None
    for selector in ['input[id*="user" i]', 'input[name*="user" i]', 'input[placeholder*="user" i]', 'input[type="text"]']:
        loc = page.locator(selector)
        if loc.count() > 0:
            username_field = loc.first
            break
            
    # Find password input
    password_field = page.locator('input[type="password"]').first
    
    # Fill details
    if username_field:
        username_field.fill(LOGIN_USER)
    password_field.fill(LOGIN_PASS)
    
    # Find submit button
    submit_btn = None
    for selector in ['input[type="submit"]', 'button', 'input[value*="Login" i]', 'input[value*="Log In" i]', 'input[id*="btn" i]']:
        loc = page.locator(selector)
        if loc.count() > 0:
            submit_btn = loc.first
            break
            
    if submit_btn:
        submit_btn.click()
        page.wait_for_load_state("networkidle")
        page.wait_for_timeout(1500)
        log("Logged in successfully.")
    else:
        log("Submit button not found!", "ERROR")

def main():
    print("=" * 80)
    print("            CGNET AUTOMATED AUDIT SCRAPER & REPORT GENERATOR")
    print("=" * 80)
    
    # Parse CLI Arguments
    today_str = datetime.now().strftime("%d %b %Y")  # e.g., "05 Aug 2026"
    parser = argparse.ArgumentParser(description="CGNET Employee Audit Scraper")
    parser.add_argument("--employee", default="Om Neupane", help="Employee name (default: Om Neupane)")
    parser.add_argument("--from-date", default=today_str, help=f"From Date, e.g. '05 Aug 2026' (default: {today_str})")
    parser.add_argument("--to-date", default=today_str, help=f"To Date, e.g. '05 Aug 2026' (default: {today_str})")
    parser.add_argument("--non-interactive", action="store_true", help="Run in full automated mode without prompt")
    args = parser.parse_args()

    employee_name = args.employee
    from_date = args.from_date
    to_date = args.to_date

    # Interactive Override
    if not args.non_interactive:
        print("Please configure the run details (Press Enter to use defaults):")
        emp_input = input(f"Employee Name [{employee_name}]: ").strip()
        if emp_input:
            employee_name = emp_input
            
        from_input = input(f"From Date [{from_date}]: ").strip()
        if from_input:
            from_date = from_input
            
        to_input = input(f"To Date [{to_date}]: ").strip()
        if to_input:
            to_date = to_input
            
    log(f"Configuration set: Employee={employee_name}, From={from_date}, To={to_date}")
    
    # Setup folders
    output_dir = os.path.dirname(os.path.abspath(__file__))
    # Format a safe filename with employee and date
    if ',' in employee_name or 'ALL TEAM' in employee_name.upper():
        safe_emp = "ALL_TEAM"
    else:
        safe_emp = re.sub(r'[^a-zA-Z0-9]', '_', employee_name)
    safe_from = re.sub(r'[^a-zA-Z0-9]', '_', from_date)
    output_file = os.path.join(output_dir, f"audit_report_{safe_emp}_{safe_from}.xlsx")

    # Start Playwright
    with sync_playwright() as p:
        log("Launching Chrome browser...")
        is_headless = os.environ.get("HEADLESS", "true").lower() != "false"
        browser = p.chromium.launch(headless=is_headless, args=["--no-sandbox", "--disable-setuid-sandbox", "--disable-dev-shm-usage", "--disable-gpu"])
        context = browser.new_context(viewport={"width": 1280, "height": 720}, ignore_https_errors=True)
        page = context.new_page()
        
        # Navigate to login portal first
        login_url = "https://billing.cgnet.com.np/h8ssrms/Login.aspx"
        target_url = "https://billing.cgnet.com.np/h8ssrms/Auditpage.aspx"
        
        log(f"Navigating to {login_url}...")
        try:
            page.goto(login_url, timeout=60000)
            safe_wait_for_networkidle(page, 10000)
        except Exception as e:
            log(f"Navigation issue: {e}", "WARNING")
            
        # Handle login
        handle_login(page)
        
        # Navigate to Audit page
        log(f"Navigating to {target_url}...")
        try:
            page.goto(target_url, timeout=30000)
            safe_wait_for_networkidle(page, 10000)
        except Exception as e:
            log(f"Could not load audit page: {e}", "ERROR")
            browser.close()
            return

        # Define JS functions for page interaction
        automate_selects_js = r"""
        ([auditFor, moduleVal, operationVal, employeeVal]) => {
            const selects = Array.from(document.querySelectorAll('select'));
            
            // 1. Audit For
            const auditForSelect = selects.find(s => 
                Array.from(s.options).some(o => o.text.trim() === auditFor || o.value.trim() === auditFor)
            );
            if (auditForSelect) {
                const opt = Array.from(auditForSelect.options).find(o => o.text.trim() === auditFor || o.value.trim() === auditFor);
                if (auditForSelect.value !== opt.value) {
                    auditForSelect.value = opt.value;
                    auditForSelect.dispatchEvent(new Event('change'));
                    return "auditFor";
                }
            }
            
            // 2. Module
            const moduleSelect = selects.find(s => 
                Array.from(s.options).some(o => o.text.trim() === moduleVal || o.value.trim() === moduleVal)
            );
            if (moduleSelect) {
                const opt = Array.from(moduleSelect.options).find(o => o.text.trim() === moduleVal || o.value.trim() === moduleVal);
                if (moduleSelect.value !== opt.value) {
                    moduleSelect.value = opt.value;
                    moduleSelect.dispatchEvent(new Event('change'));
                    return "module";
                }
            }
            
            // 3. Operation
            const operationSelect = selects.find(s => 
                Array.from(s.options).some(o => o.text.trim() === operationVal || o.value.trim() === operationVal)
            );
            if (operationSelect) {
                const opt = Array.from(operationSelect.options).find(o => o.text.trim() === operationVal || o.value.trim() === operationVal);
                if (operationSelect.value !== opt.value) {
                    operationSelect.value = opt.value;
                    operationSelect.dispatchEvent(new Event('change'));
                    return "operation";
                }
            }
            
            // 4. Employee (UserWise dropdown)
            const employeeSelect = selects.find(s => 
                Array.from(s.options).some(o => o.text.trim().toLowerCase().includes(employeeVal.toLowerCase()))
            );
            if (employeeSelect) {
                const opt = Array.from(employeeSelect.options).find(o => o.text.trim().toLowerCase().includes(employeeVal.toLowerCase()));
                if (opt && employeeSelect.value !== opt.value) {
                    employeeSelect.value = opt.value;
                    employeeSelect.dispatchEvent(new Event('change'));
                    return "employee";
                }
            }
            
            return "done";
        }
        """

        set_date_inputs_js = r"""
        ([fromDateStr, toDateStr]) => {
            const fromInput = document.getElementById('ContentPlaceHolder1_txtfrmdate') || document.querySelector('input[id*="frm" i], input[name*="frm" i]');
            const toInput = document.getElementById('ContentPlaceHolder1_txttodate') || document.querySelector('input[id*="to" i], input[name*="to" i]');
            
            if (fromInput) {
                fromInput.value = fromDateStr;
                fromInput.dispatchEvent(new Event('change'));
                fromInput.dispatchEvent(new Event('blur'));
            }
            if (toInput) {
                toInput.value = toDateStr;
                toInput.dispatchEvent(new Event('change'));
                toInput.dispatchEvent(new Event('blur'));
            }
            return !!(fromInput && toInput);
        }
        """

        click_search_js = r"""
        () => {
            const btn = document.getElementById('ContentPlaceHolder1_btnserch');
            if (btn) {
                btn.click();
                return true;
            }
            const buttons = Array.from(document.querySelectorAll('input[type="submit"], input[type="button"], button'));
            const searchBtn = buttons.find(b => {
                const val = (b.value || b.textContent || '').trim().toLowerCase();
                return val === 'search' || val.includes('search');
            });
            if (searchBtn) {
                searchBtn.click();
                return true;
            }
            return false;
        }
        """

        # Extraction loop injection scripts
        find_grid_table_js = r"""
        () => {
            const grid = document.getElementById('ContentPlaceHolder1_gdhistory') || 
                         document.querySelector('table[id*="gdhistory" i]') || 
                         document.querySelector('table.table2');
            return !!grid;
        }
        """
        
        get_rows_count_js = r"""
        () => {
            const gridTable = document.getElementById('ContentPlaceHolder1_gdhistory') || 
                              document.querySelector('table[id*="gdhistory" i]') || 
                              document.querySelector('table.table2');
            if (!gridTable) return 0;
            const rows = Array.from(gridTable.querySelectorAll('tr'));
            let count = 0;
            for (const row of rows) {
                const cells = Array.from(row.querySelectorAll('td'));
                if (cells.length < 5) continue;
                const firstCellText = cells[0].textContent.trim().toLowerCase();
                if (firstCellText === 'date') continue;
                if (row.querySelector('table') || cells.some(c => c.textContent.trim().match(/^\d+$/) && cells.length <= 2)) {
                    continue;
                }
                count++;
            }
            return count;
        }
        """
        
        get_row_data_js = r"""
        (rowIndex) => {
            const gridTable = document.getElementById('ContentPlaceHolder1_gdhistory') || 
                              document.querySelector('table[id*="gdhistory" i]') || 
                              document.querySelector('table.table2');
            if (!gridTable) return null;
            const rows = Array.from(gridTable.querySelectorAll('tr'));
            
            const dataRows = [];
            for (const row of rows) {
                const cells = Array.from(row.querySelectorAll('td'));
                if (cells.length < 5) continue;
                const firstCellText = cells[0].textContent.trim().toLowerCase();
                if (firstCellText === 'date') continue;
                if (row.querySelector('table') || cells.some(c => c.textContent.trim().match(/^\d+$/) && cells.length <= 2)) {
                    continue;
                }
                dataRows.push(row);
            }
            
            if (rowIndex >= dataRows.length) return null;
            const targetRow = dataRows[rowIndex];
            const cells = Array.from(targetRow.querySelectorAll('td'));
            
            return {
                date: cells[0].textContent.trim(),
                userName: cells[1].textContent.trim(),
                employeeName: cells[2].textContent.trim(),
                moduleName: cells[3].textContent.trim(),
                operation: cells[4].textContent.trim(),
                ipAddress: cells.length > 5 ? cells[5].textContent.trim() : '',
                remark: cells.length > 6 ? cells[6].textContent.trim() : ''
            };
        }
        """
        
        click_row_date_js = click_row_link_js = r"""
        (rowIndex) => {
            const gridTable = document.getElementById('ContentPlaceHolder1_gdhistory') || 
                              document.querySelector('table[id*="gdhistory" i]') || 
                              document.querySelector('table.table2');
            if (!gridTable) return false;
            const rows = Array.from(gridTable.querySelectorAll('tr'));
            
            const dataRows = [];
            for (const row of rows) {
                const cells = Array.from(row.querySelectorAll('td'));
                if (cells.length < 5) continue;
                const firstCellText = cells[0].textContent.trim().toLowerCase();
                if (firstCellText === 'date') continue;
                if (row.querySelector('table') || cells.some(c => c.textContent.trim().match(/^\d+$/) && cells.length <= 2)) {
                    continue;
                }
                dataRows.push(row);
            }
            
            if (rowIndex >= dataRows.length) return false;
            const cell = dataRows[rowIndex].querySelectorAll('td')[0];
            const clickTarget = cell.querySelector('a') || cell;
            clickTarget.click();
            return true;
        }
        """
        
        scroll_modal_js = r"""
        () => {
            const headers = Array.from(document.querySelectorAll('*')).filter(el => 
                el.textContent && el.textContent.trim() === 'Case Information' && el.offsetWidth > 0
            );
            if (headers.length === 0) return false;
            
            let container = null;
            let parent = headers[0].parentElement;
            while (parent && parent.tagName !== 'BODY') {
                if (parent.querySelector('input[type="button"][value="Close"]') || parent.innerText.includes('Ticket Number')) {
                    container = parent;
                    break;
                }
                parent = parent.parentElement;
            }
            
            if (!container) container = document;
            
            let scrolled = false;
            const allElements = container.querySelectorAll('*');
            for (const el of allElements) {
                if (el.scrollHeight > el.clientHeight && 
                    (window.getComputedStyle(el).overflowY === 'auto' || 
                     window.getComputedStyle(el).overflowY === 'scroll' ||
                     el.style.overflow === 'auto' ||
                     el.style.overflow === 'scroll')) {
                    el.scrollTop = el.scrollHeight;
                    scrolled = true;
                }
            }
            return scrolled;
        }
        """
        
        extract_modal_data_js = r"""
        () => {
            const targets = [
                "Title", "Ticket Number", "Account Name", "Status", "Case Origin", 
                "Customer Name", "Mobile No", "Category", "Sub Sub Category", 
                "Assigned Date", "Estimated closed time based on SLA", 
                "Estimated closed time based on TAT", "User Id", "Created Date", 
                "Last Modified Date", "Associated Location Status", "Latitude", 
                "Case Reason", "Assign Team", "Assign User", "Priority", "Type", 
                "Customer Address", "Email", "Sub Category", "Escalated Layer", 
                "SLA", "SLA Left/Total", "TAT Left/Total", "Created By", 
                "Last Modified By", "Associated Location Level", 
                "Associated Location Name", "Longitude"
            ];
            
            const headers = Array.from(document.querySelectorAll('*')).filter(el => 
                el.textContent && el.textContent.trim() === 'Case Information' && el.offsetWidth > 0
            );
            
            let container = null;
            if (headers.length > 0) {
                let parent = headers[0].parentElement;
                while (parent && parent.tagName !== 'BODY') {
                    if (parent.querySelector('input[type="button"][value="Close"]') || parent.innerText.includes('Ticket Number')) {
                        container = parent;
                        break;
                    }
                    parent = parent.parentElement;
                }
            }
            if (!container) container = document;
            
            const data = {};
            const cells = Array.from(container.querySelectorAll('td'));
            
            for (const target of targets) {
                const foundCell = cells.find(c => {
                    const text = c.textContent ? c.textContent.trim().replace(/\s+/g, ' ') : '';
                    return text === target || text === target + ':' || text === target + ' :';
                });
                
                if (foundCell) {
                    const nextCell = foundCell.nextElementSibling;
                    if (nextCell && nextCell.tagName === 'TD') {
                        data[target] = nextCell.textContent.trim();
                    } else {
                        const row = foundCell.closest('tr');
                        if (row) {
                            const rowCells = Array.from(row.querySelectorAll('td'));
                            const idx = rowCells.indexOf(foundCell);
                            if (idx !== -1 && idx + 1 < rowCells.length) {
                                data[target] = rowCells[idx + 1].textContent.trim();
                            }
                        }
                    }
                } else {
                    const divs = Array.from(container.querySelectorAll('div, span, label'));
                    const foundDiv = divs.find(d => {
                        const text = d.textContent ? d.textContent.trim().replace(/\s+/g, ' ') : '';
                        return (text === target || text === target + ':') && d.children.length === 0 && d.offsetWidth > 0;
                    });
                    if (foundDiv) {
                        let next = foundDiv.nextElementSibling;
                        if (next) {
                            data[target] = next.textContent.trim();
                        }
                    }
                }
            }

            // Extract Solution Given / Work Log / Employee Remarks from modal text
            let solutionGivenText = "";
            let fullText = (container.innerText || container.textContent || "").trim();
            
            const solutionMatch = fullText.match(/Solution Given[\s\S]*/i);
            if (solutionMatch) {
                solutionGivenText = solutionMatch[0].replace(/^Solution Given\s*/i, '').replace(/Close\s*$/i, '').trim();
            }
            
            data["Solution Given"] = solutionGivenText || "";
            data["Employee Remark / Solution Note"] = solutionGivenText || fullText;
            return data;
        }
        """
        
        close_modal_js = r"""
        () => {
            const buttons = Array.from(document.querySelectorAll('input[type="button"], button, a')).filter(el => 
                el.offsetWidth > 0
            );
            const closeBtn = buttons.find(el => {
                const val = (el.value || el.textContent || '').trim().toLowerCase();
                return val === 'close';
            });
            if (closeBtn) {
                closeBtn.click();
                return true;
            }
            return false;
        }
        """
        
        get_pagination_info_js = r"""
        () => {
            const gridTable = document.getElementById('ContentPlaceHolder1_gdhistory') || 
                              document.querySelector('table[id*="gdhistory" i]') || 
                              document.querySelector('table.table2');
            if (!gridTable) return null;
            
            let links = Array.from(gridTable.querySelectorAll('a[id*="lnbPg"], span[id*="lnbPg"]'));
            if (links.length === 0) {
                const rows = Array.from(gridTable.querySelectorAll('tr'));
                if (rows.length > 0) {
                    const lastRow = rows[rows.length - 1];
                    const candidateLinks = Array.from(lastRow.querySelectorAll('a, span'));
                    links = candidateLinks.filter(el => {
                        const txt = el.textContent.trim();
                        return /^\d+$/.test(txt) || txt === '...';
                    });
                }
            }
            
            if (links.length <= 1) return null;

            return links.map(el => {
                const text = el.textContent.trim();
                const hasHref = el.hasAttribute('href') && el.getAttribute('href').length > 0;
                const isDisabled = el.classList.contains('aspNetDisabled') || !hasHref;
                return {
                    text: text,
                    active: isDisabled,
                    clickable: !isDisabled && hasHref
                };
            }).filter(item => /^\d+$/.test(item.text) || item.text === '...');
        }
        """
        
        click_page_js = r"""
        (pageNumStr) => {
            const gridTable = document.getElementById('ContentPlaceHolder1_gdhistory') || 
                              document.querySelector('table[id*="gdhistory" i]') || 
                              document.querySelector('table.table2');
            if (!gridTable) return false;
            
            const links = Array.from(gridTable.querySelectorAll('a'));
            
            // 1. Check for exact page number
            let targetLink = links.find(el => {
                const txt = el.textContent.trim();
                return txt === String(pageNumStr) && el.hasAttribute('href') && !el.classList.contains('aspNetDisabled');
            });
            
            // 2. If page number not directly visible, click the trailing ellipsis (Next set of 10 pages)
            if (!targetLink) {
                const ellipsisLinks = links.filter(el => el.textContent.trim() === '...' && el.hasAttribute('href') && !el.classList.contains('aspNetDisabled'));
                if (ellipsisLinks.length > 0) {
                    targetLink = ellipsisLinks[ellipsisLinks.length - 1];
                }
            }
            
            if (targetLink) {
                targetLink.click();
                return true;
            }
            return false;
        }
        """

        # Prepare list of target employees
        if ',' in employee_name:
            emp_list = [e.strip() for e in employee_name.split(',') if e.strip()]
        else:
            emp_list = [employee_name.strip()]

        scraped_records = []
        
        for emp_idx, target_emp in enumerate(emp_list):
            log(f"============================================================")
            log(f"--- Starting audit scrape for Employee {emp_idx+1}/{len(emp_list)}: {target_emp} ---")
            log(f"============================================================")
            
            # Auto-configure grid dropdowns
            log(f"Auto-applying dropdown filters for '{target_emp}' (AuditFor='Employee', Module='Case', Operation='Update')...")
            for attempt in range(10):
                step = page.evaluate(automate_selects_js, ["Employee", "Case", "Update", target_emp])
                log(f"Select configuration status: {step}")
                if step == "done":
                    break
                wait_for_postback(page)
                page.wait_for_timeout(500)
                safe_wait_for_networkidle(page, 5000)
                
            # Set Dates
            log(f"Auto-setting date inputs (From: {from_date}, To: {to_date})...")
            dates_set = page.evaluate(set_date_inputs_js, [from_date, to_date])
            if not dates_set:
                log("Could not find date inputs to modify. Please adjust dates manually in browser if incorrect.", "WARNING")
            page.wait_for_timeout(500)
            
            # Click search
            log(f"Submitting query for {target_emp}...")
            searched = page.evaluate(click_search_js)
            if searched:
                wait_for_postback(page)
                page.wait_for_timeout(1500)
                safe_wait_for_networkidle(page, 8000)
            else:
                log("Search button not found. Please click Search in the browser manually.", "WARNING")
                page.wait_for_timeout(3000)

            # Safe-guard if search is slow
            try:
                page.locator('table').first.wait_for(state="visible", timeout=6000)
            except Exception:
                pass

            page_num = 1
            while True:
                log(f"[{target_emp}] Scanning page {page_num}...")
                
                has_grid = page.evaluate(find_grid_table_js)
                if not has_grid:
                    log("Grid table not found! Wait 2 seconds and check again...", "WARNING")
                    page.wait_for_timeout(2000)
                    has_grid = page.evaluate(find_grid_table_js)
                    if not has_grid:
                        log(f"No grid table found for {target_emp}.", "INFO")
                        break
                    
                rows_count = page.evaluate(get_rows_count_js)
                log(f"[{target_emp}] Found {rows_count} records on page {page_num}.")
                
                if rows_count == 0:
                    log(f"[{target_emp}] No data records found on page {page_num}.")
                    break
                    
                for i in range(rows_count):
                    row_info = page.evaluate(get_row_data_js, i)
                    if not row_info:
                        log(f"[{target_emp}] Could not extract row info for record index {i+1}", "WARNING")
                        continue
                    
                    log(f"[{target_emp}] Record {i+1}/{rows_count} (Page {page_num}): Date='{row_info['date']}' | User='{row_info['userName']}' | Employee='{row_info['employeeName']}' | Remark='{row_info['remark']}'")
                    
                    # Click row link to open modal
                    clicked = page.evaluate(click_row_link_js, i)
                    if not clicked:
                        log(f"[{target_emp}] Failed to click popup link for row {i+1}", "WARNING")
                        continue
                    
                    # Wait for ASP.NET postback
                    wait_for_postback(page)
                    page.wait_for_timeout(400)
                    
                    # Wait for modal frame to appear
                    modal_frame = None
                    for _ in range(25):
                        for frame in page.frames:
                            try:
                                has_header = frame.evaluate(r"""
                                () => {
                                    const headers = Array.from(document.querySelectorAll('*')).filter(el => 
                                        el.textContent && el.textContent.trim() === 'Case Information' && el.offsetWidth > 0
                                    );
                                    return headers.length > 0;
                                }
                                """)
                                if has_header:
                                    modal_frame = frame
                                    break
                            except Exception:
                                pass
                        if modal_frame:
                            break
                        page.wait_for_timeout(200)

                    if not modal_frame:
                        log(f"[{target_emp}] Timeout waiting for 'Case Information' popup to open inside iframe for record {i+1}", "WARNING")
                        try:
                            frame = page.frame(name="ContentPlaceHolder1_ifrm")
                            if frame:
                                frame.evaluate(close_modal_js)
                        except Exception:
                            pass
                        continue
                    
                    page.wait_for_timeout(300)
                    
                    # Scroll modal inside the frame
                    try:
                        modal_frame.evaluate(scroll_modal_js)
                    except Exception:
                        pass
                    page.wait_for_timeout(200)
                    
                    # Extract details from the frame
                    modal_data = {}
                    try:
                        modal_data = modal_frame.evaluate(extract_modal_data_js)
                    except Exception as extract_err:
                        log(f"Extraction error: {extract_err}", "ERROR")
                    
                    # Close the modal inside the frame
                    closed = False
                    try:
                        closed = modal_frame.evaluate(close_modal_js)
                    except Exception:
                        pass
                        
                    if not closed:
                        try:
                            modal_frame.locator('text=Close').first.click(timeout=2000)
                        except Exception:
                            log("Could not close the modal.", "ERROR")
                    
                    # Wait for close postback to complete
                    wait_for_postback(page)
                    page.wait_for_timeout(300)
                    
                    modal_closed = False
                    for _ in range(25):
                        try:
                            is_visible = page.locator('iframe[name="ContentPlaceHolder1_ifrm"]').is_visible()
                            if not is_visible:
                                modal_closed = True
                                break
                        except Exception:
                            modal_closed = True
                            break
                        page.wait_for_timeout(200)

                    combined_record = {
                        "Target Employee": target_emp,
                        "Grid Date": row_info["date"],
                        "Grid User Name": row_info["userName"],
                        "Grid Employee Name": row_info["employeeName"],
                        "Grid Module": row_info["moduleName"],
                        "Grid Operation": row_info["operation"],
                        "Grid IP Address": row_info["ipAddress"],
                        "Grid Remark": row_info["remark"],
                        **modal_data
                    }
                    scraped_records.append(combined_record)
                    page.wait_for_timeout(300)
                    
                # Check pagination
                pagination_items = page.evaluate(get_pagination_info_js)
                if not pagination_items:
                    log(f"[{target_emp}] No pagination elements found. Single page audit.")
                    break
                    
                active_item = next((item for item in pagination_items if item["active"]), None)
                if not active_item:
                    log(f"[{target_emp}] No active page marked. Assuming page iteration completed.", "INFO")
                    break
                    
                current_page_val = int(active_item["text"]) if active_item["text"].isdigit() else page_num
                next_page_val = current_page_val + 1
                
                # Check for direct next page number link or trailing ellipsis '...'
                has_next_number = any(item["text"] == str(next_page_val) and item["clickable"] for item in pagination_items)
                has_next_ellipsis = any(item["text"] == "..." and item["clickable"] for item in pagination_items)
                
                if has_next_number or has_next_ellipsis:
                    log(f"[{target_emp}] Navigating to next page: Page {next_page_val}...")
                    
                    first_row_before = page.evaluate(get_row_data_js, 0)
                    date_before = first_row_before["date"] if first_row_before else ""
                    
                    page.evaluate(click_page_js, str(next_page_val))
                    wait_for_postback(page)
                    page.wait_for_timeout(1500)
                    safe_wait_for_networkidle(page, 6000)
                    
                    post_pagination = page.evaluate(get_pagination_info_js)
                    new_active = next((item for item in post_pagination if item["active"]), None) if post_pagination else None
                    first_row_after = page.evaluate(get_row_data_js, 0)
                    date_after = first_row_after["date"] if first_row_after else ""
                    
                    if (new_active and new_active["text"].isdigit() and int(new_active["text"]) >= next_page_val) or (date_before != date_after):
                        page_num = int(new_active["text"]) if (new_active and new_active["text"].isdigit()) else next_page_val
                        log(f"[{target_emp}] Loaded Page {page_num} successfully.")
                    else:
                        log(f"[{target_emp}] Completed all available pages up to Page {current_page_val}.", "INFO")
                        break
                else:
                    log(f"[{target_emp}] Reached the last page (Page {current_page_val}). All pages extracted successfully.")
                    break
            
            log(f"✓ Finished scraping all audit records for {target_emp}.")
            page.wait_for_timeout(1000)
                
        log("Scraping completed. Closing browser.")
        browser.close()
        
        # --- Post-Processing & Report Generation ---
        if not scraped_records:
            log("No records were scraped. Unable to generate report.", "ERROR")
            return
            
        log(f"Scraped {len(scraped_records)} total audit records. Compiling report...")
        
        df = pd.DataFrame(scraped_records)
        
        required_cols = ['Ticket Number', 'Status', 'Category', 'Sub Category', 'Assigned Date', 'Created Date', 'Last Modified Date']
        for col in required_cols:
            if col not in df.columns:
                df[col] = None
                
        df['dt_assigned'] = df['Assigned Date'].apply(parse_date_robustly)
        df['dt_created'] = df['Created Date'].apply(parse_date_robustly)
        df['dt_completed'] = df['Last Modified Date'].apply(parse_date_robustly)
        
        df['duration_assigned_sec'] = (df['dt_completed'] - df['dt_assigned']).dt.total_seconds()
        df['duration_created_sec'] = (df['dt_completed'] - df['dt_created']).dt.total_seconds()
        
        df['duration_assigned'] = pd.to_timedelta(df['duration_assigned_sec'], unit='s')
        df['duration_created'] = pd.to_timedelta(df['duration_created_sec'], unit='s')
        df['Resolution Time (From Assigned)'] = df['duration_assigned'].apply(format_duration)
        df['Resolution Time (From Created)'] = df['duration_created'].apply(format_duration)
        
        df['Status_Cleaned'] = df['Status'].astype(str).str.strip().str.capitalize()
        
        def classify_ticket_handling(row):
            status_clean = str(row.get('Status', '')).strip().capitalize()
            raw_status = str(row.get('Status', '')).strip()
            grid_remark = str(row.get('Grid Remark', '')).strip()
            assign_team = str(row.get('Assign Team', '')).strip()
            assign_user = str(row.get('Assign User', '')).strip()
            
            remark_lower = grid_remark.lower()
            team_lower = assign_team.lower()
            user_lower = assign_user.lower()
            
            is_ms_or_transferred = (
                'ms' in team_lower or 'ms' in user_lower or
                'ms' in remark_lower or 'assign' in remark_lower or
                'transfer' in remark_lower or 'forward' in remark_lower or
                'handover' in remark_lower or 'handler' in remark_lower
            )
            
            if status_clean in ['Completed', 'Closed']:
                if is_ms_or_transferred and status_clean != 'Completed':
                    res_type = "Assigned / Transferred to MS / Other Handler"
                else:
                    res_type = f"Completed ({raw_status if raw_status else 'Completed'})"
                return True, res_type, grid_remark
            elif is_ms_or_transferred:
                res_type = "Assigned / Transferred to MS / Other Handler"
                return True, res_type, grid_remark
            else:
                res_type = f"Other / In Progress ({raw_status if raw_status else 'Unspecified'})"
                return False, res_type, grid_remark

        res_results = df.apply(classify_ticket_handling, axis=1)
        df['Is_Solved_Or_Assigned'] = [r[0] for r in res_results]
        df['Resolution Type'] = [r[1] for r in res_results]
        df['Action / Transfer Remark'] = [r[2] for r in res_results]
        
        def classify_work_type(row):
            remark = str(row.get('Grid Remark', '')).strip()
            solution = str(row.get('Solution Given', '')).strip()
            employee_note = str(row.get('Employee Remark / Solution Note', '')).strip()
            title = str(row.get('Title', '')).strip()
            category = str(row.get('Category', '')).strip()
            
            full_text = f"{title} {remark} {solution} {employee_note}".upper()
            
            # Check for 'ALCLFE' (WiFi 6 Serial Number prefix) or Old/New SN router upgrade remarks
            is_wifi6_upgrade = (
                'ALCLFE' in full_text or
                ('WIFI 6' in full_text and ('OLD' in full_text or 'NEW' in full_text or 'UPGRADE' in full_text or 'UPGARDE' in full_text or 'SN' in full_text)) or
                ('WIFI' in full_text and 'ALCL' in full_text)
            )
            
            if is_wifi6_upgrade:
                return 'WiFi 6 Upgrade'
            elif 'IPTV' in full_text or 'IPTV' in category.upper():
                return 'IPTV Issue'
            else:
                return 'General / Other Issues'

        df['Task / Issue Type'] = df.apply(classify_work_type, axis=1)
        
        # Ensure Employee Remark / Solution Note column is clean
        if 'Employee Remark / Solution Note' not in df.columns:
            df['Employee Remark / Solution Note'] = ""
            
        df['Employee Remark / Solution Note'] = df.apply(
            lambda r: str(r.get('Solution Given', '')).strip() or str(r.get('Employee Remark / Solution Note', '')).strip() or str(r.get('Grid Remark', '')).strip(),
            axis=1
        )
        
        solved_mask = df['Is_Solved_Or_Assigned']
        
        total_records = len(df)
        solved_tickets = df[solved_mask]
        num_solved = len(solved_tickets)
        
        avg_assigned_duration = solved_tickets['duration_assigned'].mean()
        avg_created_duration = solved_tickets['duration_created'].mean()
        
        # 1. Category Breakdown for Solved Tickets
        cat_group = solved_tickets.groupby(['Category', 'Sub Category'], dropna=False).agg(
            Count=('Ticket Number', 'count'),
            Avg_Duration_Assigned=('duration_assigned_sec', 'mean'),
            Avg_Duration_Created=('duration_created_sec', 'mean')
        ).reset_index()
        cat_group['Avg Resolution Time (From Assigned)'] = pd.to_timedelta(cat_group['Avg_Duration_Assigned'], unit='s').apply(format_duration)
        cat_group['Avg Resolution Time (From Created)'] = pd.to_timedelta(cat_group['Avg_Duration_Created'], unit='s').apply(format_duration)
        cat_group.drop(columns=['Avg_Duration_Assigned', 'Avg_Duration_Created'], inplace=True)
        
        # 2. TOTAL Tickets Category Breakdown (All Tickets Scraped)
        cat_total_group = df.groupby(['Category', 'Sub Category'], dropna=False).agg(
            Total_Tickets=('Ticket Number', 'count'),
            Solved_Count=('Is_Solved_Or_Assigned', 'sum')
        ).reset_index()
        cat_total_group['Solution Rate %'] = (cat_total_group['Solved_Count'] / cat_total_group['Total_Tickets'] * 100).round(1).astype(str) + '%'
        cat_total_group.rename(columns={
            'Total_Tickets': 'Total Scraped Tickets',
            'Solved_Count': 'Solved / Handled Count'
        }, inplace=True)
        
        # 3. High-level Category Summary
        cat_summary_overall = df.groupby('Category', dropna=False).agg(
            Total_Tickets=('Ticket Number', 'count'),
            Solved_Count=('Is_Solved_Or_Assigned', 'sum')
        ).reset_index()
        cat_summary_overall['Category Solution Rate %'] = (cat_summary_overall['Solved_Count'] / cat_summary_overall['Total_Tickets'] * 100).round(1).astype(str) + '%'
        cat_summary_overall.rename(columns={
            'Total_Tickets': 'Total Scraped Tickets',
            'Solved_Count': 'Solved / Handled Count'
        }, inplace=True)
        
        solved_list_cols = [
            'Ticket Number', 'Account Name', 'Category', 'Sub Category', 
            'Status', 'Resolution Type', 'Employee Remark / Solution Note', 'Action / Transfer Remark',
            'Assigned Date', 'Created Date', 'Last Modified Date',
            'Resolution Time (From Assigned)', 'Resolution Time (From Created)'
        ]
        solved_list_cols = [c for c in solved_list_cols if c in solved_tickets.columns]
        solved_list = solved_tickets[solved_list_cols].copy()
        
        summary_metrics = pd.DataFrame({
            "Metric": [
                "Total Audit Records Scraped",
                "Total Tickets Solved / Handled (Completed or MS Assigned)",
                "Ticket Solution Rate",
                "Average Completion Time (From Assigned Date)",
                "Average Completion Time (From Created Date)"
            ],
            "Value": [
                total_records,
                num_solved,
                f"{(num_solved / total_records * 100):.1f}%" if total_records > 0 else "0.0%",
                format_duration(avg_assigned_duration),
                format_duration(avg_created_duration)
            ]
        })
        
        df_export = df.drop(columns=[
            'dt_assigned', 'dt_created', 'dt_completed', 
            'duration_assigned_sec', 'duration_created_sec',
            'duration_assigned', 'duration_created', 'Status_Cleaned',
            'Is_Solved_Or_Assigned'
        ], errors='ignore')
        
        log(f"Writing Excel report to {output_file}...")
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            df_export.to_excel(writer, sheet_name="Audit Details", index=False)
            summary_metrics.to_excel(writer, sheet_name="Summary Report", index=False, startrow=0, startcol=0)
            
            start_row_overall = len(summary_metrics) + 3
            pd.DataFrame([["Overall Category Summary"]]).to_excel(writer, sheet_name="Summary Report", index=False, header=False, startrow=start_row_overall-1, startcol=0)
            cat_summary_overall.to_excel(writer, sheet_name="Summary Report", index=False, startrow=start_row_overall, startcol=0)
            
            start_row_total_cat = start_row_overall + len(cat_summary_overall) + 3
            pd.DataFrame([["Total Tickets Breakdown by Category & Sub Category"]]).to_excel(writer, sheet_name="Summary Report", index=False, header=False, startrow=start_row_total_cat-1, startcol=0)
            cat_total_group.to_excel(writer, sheet_name="Summary Report", index=False, startrow=start_row_total_cat, startcol=0)

            start_row_solved = start_row_total_cat + len(cat_total_group) + 3
            pd.DataFrame([["Detailed Solved / Handled Tickets List"]]).to_excel(writer, sheet_name="Summary Report", index=False, header=False, startrow=start_row_solved-1, startcol=0)
            solved_list.to_excel(writer, sheet_name="Summary Report", index=False, startrow=start_row_solved, startcol=0)

        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        wb = openpyxl.load_workbook(output_file)
        
        ws_details = wb["Audit Details"]
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        cell_font = Font(name="Segoe UI", size=10)
        thin_border = Border(
            left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'),
            top=Side(style='thin', color='D9D9D9'), bottom=Side(style='thin', color='D9D9D9')
        )
        
        for col in range(1, ws_details.max_column + 1):
            cell = ws_details.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
        for row in range(2, ws_details.max_row + 1):
            for col in range(1, ws_details.max_column + 1):
                cell = ws_details.cell(row=row, column=col)
                cell.font = cell_font
                cell.border = thin_border
                
        for col in ws_details.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws_details.column_dimensions[col_letter].width = min(max(max_len + 3, 10), 40)
            
        ws_summary = wb["Summary Report"]
        
        def style_summary_range(start_r, end_r, start_c, end_c, is_header=False):
            fill = PatternFill(start_color="2F5597" if is_header else "F2F2F2", end_color="2F5597" if is_header else "F2F2F2", fill_type="solid")
            font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF" if is_header else "000000")
            for r in range(start_r, end_r + 1):
                for c in range(start_c, end_c + 1):
                    cell = ws_summary.cell(row=r, column=c)
                    if is_header:
                        cell.fill = fill
                        cell.font = font
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    else:
                        cell.font = Font(name="Segoe UI", size=10)
                        cell.border = thin_border
                        if r % 2 == 1:
                            cell.fill = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")

        style_summary_range(1, 1, 1, 2, is_header=True)
        style_summary_range(2, len(summary_metrics)+1, 1, 2, is_header=False)
        
        section_title_font = Font(name="Segoe UI", size=12, bold=True, color="1F4E78")
        ws_summary.cell(row=start_row_overall, column=1).font = section_title_font
        ws_summary.cell(row=start_row_total_cat, column=1).font = section_title_font
        ws_summary.cell(row=start_row_solved, column=1).font = section_title_font
        
        style_summary_range(start_row_overall+1, start_row_overall+1, 1, len(cat_summary_overall.columns), is_header=True)
        style_summary_range(start_row_overall+2, start_row_overall+1+len(cat_summary_overall), 1, len(cat_summary_overall.columns), is_header=False)
        
        style_summary_range(start_row_total_cat+1, start_row_total_cat+1, 1, len(cat_total_group.columns), is_header=True)
        style_summary_range(start_row_total_cat+2, start_row_total_cat+1+len(cat_total_group), 1, len(cat_total_group.columns), is_header=False)
        
        style_summary_range(start_row_solved+1, start_row_solved+1, 1, len(solved_list.columns), is_header=True)
        style_summary_range(start_row_solved+2, start_row_solved+1+len(solved_list), 1, len(solved_list.columns), is_header=False)
        
        for col in ws_summary.columns:
            max_len = 0
            for cell in col:
                val = str(cell.value or '')
                if val in ["Overall Category Summary", "Total Tickets Breakdown by Category & Sub Category", "Detailed Solved / Handled Tickets List"]:
                    continue
                max_len = max(max_len, len(val))
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws_summary.column_dimensions[col_letter].width = min(max(max_len + 4, 12), 45)
            
        wb.save(output_file)
        
        # --- Print Console Summary Report ---
        print("\n" + "=" * 80)
        print("                        DAILY PERFORMANCE SUMMARY")
        print("=" * 80)
        for _, row in summary_metrics.iterrows():
            print(f"{row['Metric']:<50} : {row['Value']}")
        print("-" * 80)
        print("\nBreakdown by Ticket Category:")
        print(f"{'Category':<25} | {'Sub Category':<30} | {'Total Count':<12} | {'Solved Count'}")
        print("-" * 80)
        for _, row in cat_total_group.iterrows():
            cat = str(row['Category'])[:23]
            sub = str(row['Sub Category'])[:28]
            tot = row['Total Scraped Tickets']
            cnt = row['Solved / Handled Count']
            print(f"{cat:<25} | {sub:<30} | {tot:<12} | {cnt}")
        print("=" * 80)
        log(f"Scraped details and compiled report saved to: {output_file}")
        print("=" * 80)

if __name__ == "__main__":
    main()
