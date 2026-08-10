import pandas as pd
import re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def extract_and_categorize_tickets(csv_file='tickets.csv', output_file='filtered_tickets_categorized.xlsx'):
    df = pd.read_csv(csv_file)
    extracted_data = []

    agents_list = ["Shashikant Chaudhary", "Abhishek Poudyal", "Ajit Shrestha", "Laxman Koirala"]

    for idx, row in df.iterrows():
        popup_details = str(row['Popup Details']) if pd.notna(row['Popup Details']) else ""
        if not popup_details or popup_details == "nan" or "Error extracting details" in popup_details:
            continue

        # 1. Extract Ticket Number
        ticket_match = re.search(r'Ticket Number\s+([A-Z0-9]+)', popup_details)
        ticket_number = ticket_match.group(1) if ticket_match else ""

        # 2. Extract Status
        status_match = re.search(r'Status\s+([^\n]+?)(?:\s+Priority|\s+Case Origin|\s+Customer Name|\s+Mobile No|\s+Category|\s+Assigned Date|$)', popup_details)
        status = status_match.group(1).strip() if status_match else "Unknown"

        # 3. Extract Customer Name
        customer_match = re.search(r'Customer Name\s+([^\n]+?)(?:\s+Customer Address|\s+Account Name|\s+Mobile No)', popup_details)
        if not customer_match:
            customer_match = re.search(r'Account Name\s+([^\n]+?)(?:\s+Assign User|\s+Status)', popup_details)
        customer_name = customer_match.group(1).strip() if customer_match else ""

        # 4. Extract User ID
        userid_match = re.search(r'User Id\s+([0-9]+)', popup_details)
        user_id = userid_match.group(1) if userid_match else ""
        if not user_id:
            mobile_match = re.search(r'Mobile No\s+([0-9]+)', popup_details)
            user_id = mobile_match.group(1) if mobile_match else ""

        # 5. Extract Title
        title_match = re.search(r'Title\s+([^\n]+?)(?:\s+Case Reason|\s+Assign Team|\s+Account Name)', popup_details)
        title = title_match.group(1).strip() if title_match else ""

        # 6. Extract Category & Sub Category
        category_match = re.search(r'Category\s+([^\n]+?)(?:\s+Sub Category|\s+Sub Sub Category)', popup_details)
        category = category_match.group(1).strip() if category_match else ""

        subcategory_match = re.search(r'Sub Category\s+([^\n]+?)(?:\s+Sub Sub Category|\s+Assigned Date)', popup_details)
        subcategory = subcategory_match.group(1).strip() if subcategory_match else ""

        # 7. Extract Dates
        created_date_match = re.search(r'Created Date\s+([^\n]+?)(?:\s+Last Modified|\s+Associated Location)', popup_details)
        created_date = created_date_match.group(1).strip() if created_date_match else ""

        assigned_date = ""
        laxman_pattern = r'I:[A-Za-z\s]+\(([^)]+)\)\s*:\s*Case Assign to Laxman Koirala'
        laxman_match = re.search(laxman_pattern, popup_details)
        if laxman_match:
            assigned_date = laxman_match.group(1).strip()

        completed_date = ""
        agent = ""
        for name in agents_list:
            completed_pattern = rf'I:{re.escape(name)}\s*\(([^)]+)\)\s*:\s*\n\s*Change Status:[^\n]*To (?:Completed|Closed)'
            completed_match = re.search(completed_pattern, popup_details, re.DOTALL)
            if completed_match:
                completed_date = completed_match.group(1).strip()
                agent = name
                break

        # Fallback assign user if agent not found from status change
        if not agent:
            assign_user_match = re.search(r'Assign User\s+([^\n]+)', popup_details)
            if assign_user_match:
                agent = assign_user_match.group(1).strip()

        # 8. Extract Solution / Remarks
        solution_section = ""
        for name in agents_list:
            solution_pattern = rf'Solution Given\s*\n\s*{re.escape(name)}\s*\([^)]+\)\s*:\s*\n\s*(.+?)(?:\n\s*[A-Za-z]+|$)'
            solution_match = re.search(solution_pattern, popup_details, re.DOTALL)
            if solution_match:
                solution_section = solution_match.group(1).strip()
                break

        if not solution_section:
            rem_match = re.search(r'Previous Remark\s*\n\s*(.+?)(?:\n\n|\Z)', popup_details, re.DOTALL)
            if rem_match:
                solution_section = rem_match.group(1).strip()[:300]

        # 9. WiFi 6 Detection
        wifi6_regex = r'wifi\s*6|wifi-6|wifi6|router\s+6|6g\s+router|upgrade\s+to\s+wifi\s*6'
        is_wifi6 = bool(re.search(wifi6_regex, popup_details, re.IGNORECASE))

        extracted_data.append({
            'Ticket Number': ticket_number,
            'Status': status,
            'Customer Name': customer_name,
            'User ID': user_id,
            'Title': title,
            'Category': category,
            'Sub Category': subcategory,
            'Created Date': created_date,
            'Assigned Date': assigned_date,
            'Completed Date': completed_date,
            'Agent': agent,
            'Remarks / Solution': solution_section,
            'Is WiFi 6': is_wifi6
        })

    all_df = pd.DataFrame(extracted_data).drop_duplicates(subset=['Ticket Number'])
    print(f"Extracted total unique tickets: {len(all_df)}")

    # DataFrames per Category
    completed_closed_df = all_df[all_df['Status'].isin(['Completed', 'Closed'])].copy()
    on_hold_df = all_df[all_df['Status'].isin(['On Hold', 'Hold'])].copy()
    in_progress_df = all_df[all_df['Status'].isin(['In Progress', 'Progress'])].copy()
    
    wifi6_completed_df = all_df[
        all_df['Status'].isin(['Completed', 'Closed']) & (all_df['Is WiFi 6'] == True)
    ].copy()

    # Drop helper column for output tabs
    cols_to_display = [
        'Ticket Number', 'Status', 'Customer Name', 'User ID', 'Title',
        'Category', 'Sub Category', 'Created Date', 'Assigned Date',
        'Completed Date', 'Agent', 'Remarks / Solution'
    ]

    # Build Summary DataFrame
    total_count = len(all_df)
    summary_data = [
        {'Category': 'Total Scraped Tickets', 'Count': total_count, 'Percentage': '100%'},
        {'Category': 'Completed & Closed Tickets', 'Count': len(completed_closed_df), 'Percentage': f"{(len(completed_closed_df)/total_count*100):.1f}%" if total_count > 0 else '0%'},
        {'Category': 'On Hold Tickets', 'Count': len(on_hold_df), 'Percentage': f"{(len(on_hold_df)/total_count*100):.1f}%" if total_count > 0 else '0%'},
        {'Category': 'In Progress Tickets', 'Count': len(in_progress_df), 'Percentage': f"{(len(in_progress_df)/total_count*100):.1f}%" if total_count > 0 else '0%'},
        {'Category': 'WiFi 6 Completed Tickets', 'Count': len(wifi6_completed_df), 'Percentage': f"{(len(wifi6_completed_df)/total_count*100):.1f}%" if total_count > 0 else '0%'}
    ]
    summary_df = pd.DataFrame(summary_data)

    # Save to Multi-Tab Excel Workbook
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        summary_df.to_excel(writer, sheet_name='Summary', index=False)
        completed_closed_df[cols_to_display].to_excel(writer, sheet_name='Completed & Closed', index=False)
        on_hold_df[cols_to_display].to_excel(writer, sheet_name='On Hold', index=False)
        in_progress_df[cols_to_display].to_excel(writer, sheet_name='In Progress', index=False)
        wifi6_completed_df[cols_to_display].to_excel(writer, sheet_name='WiFi 6 Completed', index=False)

    # Apply openpyxl Styling
    wb = openpyxl.load_workbook(output_file)

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    data_font = Font(name="Calibri", size=10)

    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )

    for sheetname in wb.sheetnames:
        ws = wb[sheetname]
        ws.views.sheetView[0].showGridLines = True

        # Header style
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Data cell style & auto width calculation
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.font = data_font
                cell.border = thin_border
                cell.alignment = Alignment(vertical="top")

        # Auto-adjust column widths
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val_str = str(cell.value or '')
                if len(val_str) > max_len:
                    max_len = len(val_str)
            ws.column_dimensions[col_letter].width = min(max(max_len + 4, 12), 50)

    wb.save(output_file)
    wb.save('filtered_tickets_agents.xlsx') # Keep updated copy in standard file as well

    print("\n[OK] Processing Complete!")
    print(f"[OK] Summary Breakdown:\n{summary_df.to_string(index=False)}")
    print(f"\n[OK] Saved multi-tab report to '{output_file}' and 'filtered_tickets_agents.xlsx'")

if __name__ == '__main__':
    extract_and_categorize_tickets()
